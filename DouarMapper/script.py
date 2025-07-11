# Helper function for confirmation
skip_all = False

def confirm_action(prompt):
    global skip_all
    while True:
        answer = input(prompt).strip().lower()
        if answer == "skipall":
            skip_all = True
            return "skipall"
        if answer in ("y", "n"):
            sure = input("Are you sure? (y/n): ").strip().lower()
            if sure == "y":
                return answer
            # If not sure, re-prompt
        # If not y/n, re-prompt

def confirm_selection(prompt, valid_indices):
    global skip_all
    while True:
        answer = input(prompt).strip()
        if answer.lower() == "skipall":
            skip_all = True
            return "skipall"
        try:
            choice = int(answer)
            if choice in valid_indices:
                sure = input("Are you sure? (y/n): ").strip().lower()
                if sure == "y":
                    return choice
                # If not sure, re-prompt
        except Exception:
            pass  # re-prompt

import openpyxl
from copy import copy
import unicodedata
from openpyxl import Workbook
from difflib import SequenceMatcher

# Part 1 :
# This part Handles the position of the table in the sheet :

# Load workbook and sheet
wb = openpyxl.load_workbook("New Douars.xlsx")
ws = wb.active

# Get user input for table boundaries
print("Enter the starting row of the table:")
start_row = int(input())
print("Enter the ending row of the table:")
end_row = int(input())
print("Enter the starting column of the table  (A=1, B=2, C=3 ...):")
start_col = int(input())
print("Enter the ending column of the table  (A=1, B=2, C=3 ...):")
end_col = int(input())

row_offset = start_row - 1
col_offset = start_col - 1

merged_ranges = []

for merged_cell in ws.merged_cells.ranges:
    min_row, min_col, max_row, max_col = merged_cell.min_row, merged_cell.min_col, merged_cell.max_row, merged_cell.max_col
    if (start_row <= max_row and end_row >= min_row) and (start_col <= max_col and end_col >= min_col):
        merged_ranges.append((min_row, min_col, max_row, max_col))

for merge in merged_ranges:
    ws.unmerge_cells(start_row=merge[0], start_column=merge[1], end_row=merge[2], end_column=merge[3])


for row in range(start_row, end_row + 1):
    for col in range(start_col, end_col + 1):
        source_cell = ws.cell(row=row, column=col)
        target_cell = ws.cell(row=row - row_offset, column=col - col_offset)

        target_cell.value = source_cell.value
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = copy(source_cell.number_format)
        target_cell.border = copy(source_cell.border)

        source_cell.value = None
        source_cell.font = openpyxl.styles.Font()
        source_cell.fill = openpyxl.styles.PatternFill()
        source_cell.alignment = openpyxl.styles.Alignment()
        source_cell.number_format = 'General'
        source_cell.border = openpyxl.styles.Border()


for merge in merged_ranges:
    min_row, min_col, max_row, max_col = merge
    new_min_row = min_row - row_offset
    new_max_row = max_row - row_offset
    new_min_col = min_col - col_offset
    new_max_col = max_col - col_offset

    ws.merge_cells(start_row=new_min_row, start_column=new_min_col,
                   end_row=new_max_row, end_column=new_max_col)



# Part 2.1: Fill down the first column safely (to handle merged-like blanks without crashing)

last_value = None
for row in range(start_row, end_row + 1):
    cell = ws.cell(row=row, column=start_col)  # <-- original cell position, not the moved one

    if cell.value is not None and str(cell.value).strip() != "":
        last_value = cell.value
    else:
        if last_value is not None:
            cell.value = last_value


# Part 3: Generate dictionary of committees and their unique douars (case-insensitive merge)

committee_col = 1  # Column 1 (A) for Committee
douar_col = 2      # Column 2 (B) for Douar

douars_by_committee = {}
original_committee_names = {}  # lowercased: original

# Find the last row with data in the new table
max_row = ws.max_row
for row in range(1, max_row + 1):
    committee = ws.cell(row=row, column=committee_col).value
    douar = ws.cell(row=row, column=douar_col).value
    if committee is None or douar is None:
        continue
    committee = str(committee).strip()
    douar = str(douar).strip()
    if committee == '' or douar == '':
        continue
    committee_key = committee.casefold()
    if committee_key not in original_committee_names:
        original_committee_names[committee_key] = committee  # preserve first encountered casing
    if committee_key not in douars_by_committee:
        douars_by_committee[committee_key] = set()
    douars_by_committee[committee_key].add(douar)

# Convert sets to sorted lists for output
final_dict = {}
for committee_key, douars_set in douars_by_committee.items():
    original_name = original_committee_names[committee_key]
    final_dict[original_name] = sorted(list(douars_set))

# Sort douars for each committee at the very beginning
for committee in final_dict:
    final_dict[committee] = sorted(final_dict[committee])

print("\nDictionary of Committees and their Douars (case-insensitive merge):")
for committee, douars in final_dict.items():
    print(f"{committee}: {douars}")

# --- Fuzzy grouping and merging for committee names ---
# (No deduplication or user prompt for douars)

def normalize_name(name):
    # Remove accents, lowercase, strip, collapse spaces
    name = unicodedata.normalize('NFKD', name)
    name = ''.join([c for c in name if not unicodedata.combining(c)])
    name = name.lower()
    name = ' '.join(name.split())
    return name

def similarity(a, b):
    return SequenceMatcher(None, a, b).ratio()

# Group similar committee names
SIMILARITY_THRESHOLD = 0.8  # 80%

committee_names = list(final_dict.keys())
groups = []
used = set()

for i, name in enumerate(committee_names):
    if name in used:
        continue
    group = [name]
    used.add(name)
    norm_name = normalize_name(name)
    for j in range(i+1, len(committee_names)):
        other = committee_names[j]
        if other in used:
            continue
        norm_other = normalize_name(other)
        if similarity(norm_name, norm_other) > SIMILARITY_THRESHOLD:
            group.append(other)
            used.add(other)
    if len(group) > 1:
        groups.append(group)

# Prompt user to confirm if similar names are the same before merging
name_map = {name: name for name in committee_names}  # default: self
for group in groups:
    print(f"\nThese committee names are similar (>80%):")
    for idx, name in enumerate(group):
        print(f"  {idx+1}. {name}")
    confirm = confirm_action("Do these committee names refer to the same entity? (y/n): ")
    if confirm == "skipall":
        print("\nSkipAll detected. Skipping all further deduplication and saving results...")
        break
    if confirm == "y":
        valid_indices = list(range(1, len(group)+1))
        choice = confirm_selection(f"Choose the canonical name for this group (1-{len(group)}): ", valid_indices)
        if choice == "skipall":
            print("\nSkipAll detected. Skipping all further deduplication and saving results...")
            break
        canonical = group[choice-1]
        for name in group:
            name_map[name] = canonical
    # If 'n', do nothing (keep as separate)
if skip_all:
    # Save and exit early
    wb_final = Workbook()
    ws_final = wb_final.active
    ws_final.title = "Committees and Douars"
    ws_final.append(["Committee", "Douar"])
    for committee, douars in douars_by_canonical.items():
        for douar in douars:
            ws_final.append([committee, douar])
    wb_final.save("Douar_Final.xlsx")
    print("\nSaved the final committee-douar mapping to Douar_Final.xlsx (SkipAll mode)")
    exit()

# Merge douars under the chosen canonical names
douars_by_canonical = {}
for name, douars in final_dict.items():
    canonical = name_map[name]
    if canonical not in douars_by_canonical:
        douars_by_canonical[canonical] = set()
    douars_by_canonical[canonical].update(douars)

# For each committee, lowercase all douar names and remove exact duplicates automatically
for committee in list(douars_by_canonical.keys()):
    douars = douars_by_canonical[committee]
    seen = set()
    unique_douars = []
    original_map = {}  # lowercased: original (first encountered)
    for d in douars:
        d_lower = d.lower()
        if d_lower not in seen:
            unique_douars.append(d_lower)
            seen.add(d_lower)
            original_map[d_lower] = d  # preserve original casing for display
    douars_by_canonical[committee] = unique_douars
    douars_by_canonical[committee + '_original_map'] = original_map


# For each committee, check for >80% similarity among unique lowercased douar names and ask the user
for committee in list(douars_by_canonical.keys()):
    if committee.endswith('_original_map'):
        continue
    douar_list = douars_by_canonical[committee]
    original_map = douars_by_canonical.get(committee + '_original_map', {})
    # Iteratively group and resolve similar douars
    douars_to_remove = set()
    working_list = douar_list[:]
    while True:
        # Find all groups of >80% similar names
        douar_groups = []
        used = set()
        for i, name in enumerate(working_list):
            if name in used:
                continue
            group = [name]
            used.add(name)
            norm_name = normalize_name(name)
            for j in range(i+1, len(working_list)):
                other = working_list[j]
                if other in used:
                    continue
                norm_other = normalize_name(other)
                if similarity(norm_name, norm_other) > SIMILARITY_THRESHOLD:
                    group.append(other)
                    used.add(other)
            if len(group) > 1:
                douar_groups.append(group)
        if not douar_groups or skip_all:
            break
        for group in douar_groups:
            # Only consider names not already marked for removal
            group = [name for name in group if name not in douars_to_remove]
            if len(group) < 2:
                continue
            print(f"\nIn committee '{committee}', these douar names are similar (>80%):")
            for idx, name in enumerate(group):
                print(f"  {idx+1}. {original_map.get(name, name)}")
            if len(group) == 2:
                confirm = confirm_action("Do these douar names refer to the same entity? (y/n): ")
                if confirm == "skipall":
                    skip_all = True
                    print("\nSkipAll detected. Skipping all further deduplication and saving results...")
                    break
                if confirm == "y":
                    valid_indices = [1, 2]
                    choice = confirm_selection(f"Choose the douar name to keep for this group (1-2): ", valid_indices)
                    if choice == "skipall":
                        skip_all = True
                        print("\nSkipAll detected. Skipping all further deduplication and saving results...")
                        break
                    to_keep = group[choice-1]
                    for name in group:
                        if name != to_keep:
                            douars_to_remove.add(name)
                # If 'n', do nothing (keep all)
            else:
                # More than two names: let user select multiple sets of matches
                selection = input("If some of these are the same, enter their numbers as groups separated by spaces (e.g. 1,5 8,4 2,3,9), or type 'no' if none match: ").strip().lower()
                if selection.lower() == "skipall":
                    skip_all = True
                    print("\nSkipAll detected. Skipping all further deduplication and saving results...")
                    break
                if selection != "no":
                    try:
                        sets = [s for s in selection.split() if s]
                        for s in sets:
                            indices = [int(x)-1 for x in s.split(',') if x.strip().isdigit() and 0 <= int(x)-1 < len(group)]
                            if len(indices) > 1:
                                print("You selected:")
                                for idx in indices:
                                    print(f"  {idx+1}. {original_map.get(group[idx], group[idx])}")
                                valid_indices = [i+1 for i in indices]
                                choice = confirm_selection(f"Choose the douar name to keep for this selection (choose one of {','.join(str(i) for i in valid_indices)}): ", valid_indices)
                                if choice == "skipall":
                                    skip_all = True
                                    print("\nSkipAll detected. Skipping all further deduplication and saving results...")
                                    break
                                to_keep = group[choice-1]
                                for idx in indices:
                                    if group[idx] != to_keep:
                                        douars_to_remove.add(group[idx])
                    except Exception:
                        pass  # If input is invalid, do nothing (keep all)
        if skip_all:
            break
        # Update working_list to exclude removed douars
        working_list = [d for d in working_list if d not in douars_to_remove]
    douars_by_canonical[committee] = [original_map[d] for d in working_list]
    del douars_by_canonical[committee + '_original_map']
    if skip_all:
        break

if skip_all:
    # Sort douars for each committee before saving (case-insensitive)
    for committee in douars_by_canonical:
        douars_by_canonical[committee] = sorted(douars_by_canonical[committee], key=lambda x: x.lower())
    wb_final = Workbook()
    ws_final = wb_final.active
    ws_final.title = "Committees and Douars"
    ws_final.append(["Committee", "Douar"])
    for committee, douars in douars_by_canonical.items():
        for douar in douars:
            ws_final.append([committee, douar])
    wb_final.save("Douar_Final.xlsx")
    print("\nSaved the final committee-douar mapping to Douar_Final.xlsx (SkipAll mode)")
    exit()

# Print the final merged dictionary (after lowercasing, exact deduplication, and fuzzy user-confirmed deduplication)
print("\nFinal dictionary of Committees and their Douars (after lowercasing, exact and fuzzy deduplication):")
for committee in list(douars_by_canonical.keys()):
    douars = douars_by_canonical[committee]
    print(f"{committee}: {douars}")

# Sort douars for each committee before saving (case-insensitive)
for committee in douars_by_canonical:
    douars_by_canonical[committee] = sorted(douars_by_canonical[committee], key=lambda x: x.lower())

# Remove >90% similar douars within each committee (fuzzy deduplication)
for committee in douars_by_canonical:
    douar_list = douars_by_canonical[committee]
    unique_douars = []
    for i, d1 in enumerate(douar_list):
        is_duplicate = False
        for d2 in unique_douars:
            if similarity(normalize_name(d1), normalize_name(d2)) > 0.9:
                is_duplicate = True
                break
        if not is_duplicate:
            unique_douars.append(d1)
    douars_by_canonical[committee] = unique_douars

# Assign unique sequential IDs to every douar (across all committees)
douar_rows = []
for committee, douars in douars_by_canonical.items():
    for douar in douars:
        douar_rows.append((committee, douar))

# Generate IDs: H01, H02, ...
ids = [f"H{str(i+1).zfill(2)}" for i in range(len(douar_rows))]

# Save the final dictionary to a new Excel file 'Douar_Final.xlsx' with Committee, Douar, ID
wb_final = Workbook()
ws_final = wb_final.active
ws_final.title = "Committees and Douars"
ws_final.append(["Committee", "Douar", "ID"])
for (committee, douar), id_ in zip(douar_rows, ids):
    ws_final.append([committee, douar, id_])
wb_final.save("Douar_Final.xlsx")
print("\nSaved the final committee-douar mapping to Douar_Final.xlsx (with IDs)")

